import requests
from bs4 import BeautifulSoup
import xlsxwriter
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

# Base URL for requests
base_url = "https://www.orsr.sk/hladaj_subjekt.asp"

# Function to get the record count
def get_record_count(query):
    response = requests.get(query)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find the table with the required information
    table = soup.find("table", {"border": "0", "cellpadding": "10", "cellspacing": "2", "align": "center"})
    if not table:
        return None

    # Look for the row with the record count
    text = table.find_all("td")[1].get_text()
    start_str = "Záznamy:"
    start_idx = text.find(start_str)
    if start_idx == -1:
        return None

    end_idx = text.find(" ", start_idx + len(start_str))

    # Extract the record count
    record_count_str = text[start_idx + len(start_str):end_idx].split('/')[1].replace('\u00a0', '').strip()

    return int(record_count_str)

# Function to process combinations starting with a specific letter
def process_letter(letter):
    results = []
    letters = 'abcdefghijklmnopqrstuvwxyz'

    for second_letter in letters:
        combo = f"{letter}{second_letter}"
        query = f"{base_url}?OBMENO={combo}&PF=0&SID=0&S=&R=on&STR=1"
        print(f"Checking combination: {combo}")
        record_count = get_record_count(query)

        if record_count is None:
            print(f"No results for combination: {combo}")
            continue

        print(f"Records for {combo}: {record_count}")

        if record_count <= 500:
            results.append((combo, 0, 0, record_count))
        else:
            pf_options = [1, 2, 17, 3, 4, 5, 6, 7, 20, 19, 16, 8, 11, 14, 15]
            sid_options = range(2, 10)
            for pf in pf_options:
                for sid in sid_options:
                    query = f"{base_url}?OBMENO={combo}&PF={pf}&SID={sid}&S=&R=on&STR=1"
                    #print(f"Checking combination: {combo} with PF={pf} and SID={sid}")
                    record_count = get_record_count(query)

                    if record_count is None:
                        print(f"No results for combination: {combo} with PF={pf} and SID={sid}")
                        continue

                    print(f"Records for {combo} with PF={pf} and SID={sid}: {record_count}")

                    if record_count <= 500:
                        results.append((combo, pf, sid, record_count))
                    else:
                        for third_letter in letters:
                            sub_combo = f"{combo}{third_letter}"
                            sub_query = f"{base_url}?OBMENO={sub_combo}&PF={pf}&SID={sid}&S=&R=on&STR=1"
                            sub_record_count = get_record_count(sub_query)
                            if sub_record_count is not None:
                                print(f"Records for {sub_combo} with PF={pf} and SID={sid}: {sub_record_count}")
                                if sub_record_count <= 500:
                                    results.append((sub_combo, pf, sid, sub_record_count))
                                else:
                                    for fourth_letter in letters:
                                        final_combo = f"{sub_combo}{fourth_letter}"
                                        final_query = f"{base_url}?OBMENO={final_combo}&PF={pf}&SID={sid}&S=&R=on&STR=1"
                                        final_record_count = get_record_count(final_query)
                                        if final_record_count is not None:
                                            print(f"Records for {final_combo} with PF={pf} and SID={sid}: {final_record_count}")
                                            if final_record_count <= 500:
                                                results.append((final_combo, pf, sid, final_record_count))
                                            else:
                                                for fifth_letter in letters:
                                                    ultimate_combo = f"{final_combo}{fifth_letter}"
                                                    ultimate_query = f"{base_url}?OBMENO={ultimate_combo}&PF={pf}&SID={sid}&S=&R=on&STR=1"
                                                    ultimate_record_count = get_record_count(ultimate_query)
                                                    if ultimate_record_count is not None:
                                                        print(f"Records for {ultimate_combo} with PF={pf} and SID={sid}: {ultimate_record_count}")
                                                        if ultimate_record_count <= 500:
                                                            results.append((ultimate_combo, pf, sid, ultimate_record_count))
    
    return results
   

# Start time measurement
start_time = time.time()

letters = 'abcdefghijklmnopqrstuvwxyz'
all_results = []

# Using multithreading to speed up the process
with ThreadPoolExecutor(max_workers=len(letters)) as executor:
    future_to_letter = {executor.submit(process_letter, letter): letter for letter in letters}
    for future in as_completed(future_to_letter):
        letter = future_to_letter[future]
        try:
            results = future.result()
            all_results.extend(results)
        except Exception as exc:
            print(f"Error occurred: {letter} - {exc}")

# Filter results to keep only those with record count not exceeding 500
filtered_results = [result for result in all_results if result[3] <= 500]

# Save initial results to an Excel file
workbook = xlsxwriter.Workbook('initial_results.xlsx')
worksheet = workbook.add_worksheet()

worksheet.write('A1', 'Combination')
worksheet.write('B1', 'PF')
worksheet.write('C1', 'SID')
worksheet.write('D1', 'Record Count')

row = 1
for combo, pf, sid, count in filtered_results:
    worksheet.write(row, 0, combo)
    worksheet.write(row, 1, pf)
    worksheet.write(row, 2, sid)
    worksheet.write(row, 3, count)
    row += 1

workbook.close()
print("Initial stage completed. Results saved to initial_results.xlsx")

# Function to collect data with a given combination of PF and SID
def collect_data(combo, pf, sid):
    records = []
    page = 1
    total_records_collected = 0
    query = f"{base_url}?OBMENO={combo}&PF={pf}&SID={sid}&S=&R=on&STR={{}}"

    while True:
        response = requests.get(query.format(page))
        soup = BeautifulSoup(response.content, 'html.parser')

        # Extracting the record count and pages
        info_table = soup.find("table", {"border": "0", "cellpadding": "10", "cellspacing": "2", "align": "center"})
        if info_table:
            info_text = info_table.find_all("td")[1].get_text()
            start_idx = info_text.find("Záznamy:")
            if start_idx != -1:
                info_text = info_text[start_idx:]
                end_idx = info_text.find(" ")
                if end_idx != -1:
                    info_text = info_text[len("Záznamy:"):end_idx].split('/')
                    current_records, total_records = map(int, info_text[0].strip().split('-')[1]), int(info_text[1].strip().replace('\u00a0', ''))
                    if total_records_collected >= total_records:
                        break

        rows = soup.find_all('tr', bgcolor=["#EEEEEE", "#DDDDDD"])
        if not rows:
            break

        for row in rows:
            cells = row.find_all('td')
            if len(cells) < 2:
                continue

            # Extracting the company name and links to its pages
            company_name = cells[1].get_text(strip=True)
            company_link = cells[1].find('a')['href']
            full_link = cells[2].find_all('a')[1]['href']
            collection_link = cells[3].find('a')['href']
            records.append((company_name, company_link, full_link, collection_link))
            total_records_collected += 1

        print(f"Page {page} for combination {combo} with PF={pf} and SID={sid}: found {len(rows)} records")

        # Check if further collection is needed
        if len(rows) < 20 or page >= 25:
            break

        page += 1

    return records

# Load initial results from the Excel file
workbook = openpyxl.load_workbook('initial_results.xlsx')
sheet = workbook.active

# Create a new Excel file to save final results
workbook_final = xlsxwriter.Workbook('final_results.xlsx')
worksheet_final = workbook_final.add_worksheet()

worksheet_final.write('A1', 'Company Name')
worksheet_final.write('B1', 'First Link')
worksheet_final.write('C1', 'Second Link')
worksheet_final.write('D1', 'Third Link')

row_final = 1

# Function to process each row from the initial results
def process_row(row):
    combo, pf, sid, count = row
    if count > 0:
        # Collect data for each combination
        records = collect_data(combo, pf, sid)
        print(f"Collected {len(records)} records for combination: {combo}, PF={pf}, SID={sid}")
        return [(name, link, full_link, collection_link) for name, link, full_link, collection_link in records]
    return []

# Using multithreading to speed up the data collection process
with ThreadPoolExecutor(max_workers=10) as executor:
    future_to_row = {executor.submit(process_row, row): row for row in sheet.iter_rows(min_row=2, values_only=True)}
    total_records_saved = 0

    for future in as_completed(future_to_row):
        try:
            records = future.result()
            for record in records:
                worksheet_final.write(row_final, 0, record[0])
                worksheet_final.write(row_final, 1, record[1])
                worksheet_final.write(row_final, 2, record[2])
                worksheet_final.write(row_final, 3, record[3])
                row_final += 1
                total_records_saved += 1
        except Exception as exc:
            print(f"Error occurred while processing row: {exc}")

workbook_final.close()
print(f"Final stage completed. Saved {total_records_saved} records to final_results.xlsx")

# End time measurement
end_time = time.time()
elapsed_time = end_time - start_time
print(f"Total execution time: {elapsed_time:.2f} seconds")
